"""Generate a clean, self-contained rebuild of the SQ-LNS Allocation Tool.

    python tools/build_clean_workbook.py

Reads site/data/base.json (produced by build_data.py) and writes
dist/SQ-LNS Allocation Tool (rebuilt).xlsx.

Design
------
The original is 45 sheets and ~495 MB of sheet XML. Three things drive that
bulk, and each has a structural fix:

1. **Eight strategy sheets.** Each physically re-sorts all 9,684 rows so a
   running total can walk down them in priority order. That is unnecessary.
   While the pool is unexhausted the greedy fill gives every geography its full
   remaining need, and afterwards it gives zero, so

       cumulative allocated above = MIN(cumulative need above, pool)

   and therefore

       allocated = MIN(need, MAX(0, pool - cumulative need above))

   which needs no sorting, just a rank column and a SUMIFS. Eight sheets become
   three columns. This is exactly equivalent, not an approximation.

2. **Duplicated allocation and quantification calc sheets.** The originals are
   identical except for which input cells they read. Here one model sheet per
   level serves both, and quantification reads it through its own scaling.

3. **Twelve dormant sheets** from previous ANRiN/UNICEF exercises, reachable
   only through nine stale cells. Dropped.

The rebuild also avoids every Google-only function (`COUNTUNIQUEIFS`, `SORT`,
`SPARKLINE`) and every dynamic-array formula, so it recalculates in Excel 2016+
and in Google Sheets. Ranks are precomputed here because they depend only on
static source data, never on user inputs.
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "site", "data", "base.json")
OUT_DIR = os.path.join(ROOT, "dist")
OUT = os.path.join(OUT_DIR, "SQ-LNS Allocation Tool (rebuilt).xlsx")

# ------------------------------------------------------------------ styling

INK = "0B0B0B"
MUTED = "898781"
ACCENT = "2A78D6"
HEAD_FILL = PatternFill("solid", fgColor="1F3B63")
BAND_FILL = PatternFill("solid", fgColor="EAF2FD")
INPUT_FILL = PatternFill("solid", fgColor="FFF6DC")
TITLE = Font(name="Calibri", size=15, bold=True, color=INK)
H2 = Font(name="Calibri", size=11, bold=True, color=ACCENT)
HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10, color=INK)
SMALL = Font(name="Calibri", size=9, color=MUTED)
BOLD = Font(name="Calibri", size=10, bold=True, color=INK)
THIN = Side(style="thin", color="D8D8D2")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

N0 = "#,##0"
N1 = "#,##0.0"
N2 = "#,##0.00"
PCT = "0.0%"
MONEY = '"$"#,##0'
NGN = '"NGN "#,##0'


def header_row(ws, row, labels, widths=None, freeze=None):
    for i, text in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=text)
        c.font = HEAD
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gcl(i)].width = w
    if freeze:
        ws.freeze_panes = freeze


def label(ws, cell, text, font=BODY):
    ws[cell] = text
    ws[cell].font = font


def input_cell(ws, cell, value, fmt=None):
    ws[cell] = value
    ws[cell].fill = INPUT_FILL
    ws[cell].border = BOX
    ws[cell].font = BOLD
    if fmt:
        ws[cell].number_format = fmt


# ------------------------------------------------------------------- ranking


def compute_ranks(rows, risk_order):
    """Priority ranks per strategy. These depend only on static source data.

    Ties break exactly as the original's SORT() calls do, and the source row
    index is the final tiebreak so the order is deterministic.
    """
    idx = list(range(len(rows)))

    mort = sorted(idx, key=lambda i: (-rows[i]["u5mr"], -rows[i]["stunting"], i))
    stunt = sorted(idx, key=lambda i: (-rows[i]["stunting"], -rows[i]["u5mr"], i))
    thresh = sorted(idx, key=lambda i: (risk_order[i], -rows[i]["u5mr"], i))

    ranks = [{} for _ in rows]
    for rank, i in enumerate(mort, 1):
        ranks[i]["mortality"] = rank
    for rank, i in enumerate(stunt, 1):
        ranks[i]["stunting"] = rank
    for rank, i in enumerate(thresh, 1):
        ranks[i]["threshold"] = rank
    return ranks


def risk_sort_key(row, thresholds):
    """Position in the risk ordering, 0 worst. Mirrors the workbook's rule.

    Reproduces the source defect for level 1.3, whose stunting criterion is lost
    to a dangling reference. Documented on the README sheet of the output.
    """
    for pos, t in enumerate(thresholds):
        stunting_met = True if t["level"] == "1.3" else row["stunting"] >= t["stunting"]
        if row["u5mr"] >= t["u5mr"] and (stunting_met or row["wasting"] >= t["wasting"]):
            return pos
    return len(thresholds)


# -------------------------------------------------------------- model sheets


def build_model_sheet(wb, name, rows, level, constants, zones):
    """One row per geography: static data, derived columns, all four strategies."""
    ward = level == "wards"
    ws = wb.create_sheet(name)

    thresholds = constants["riskThresholds"]
    order = [risk_sort_key(r, thresholds) for r in rows]
    ranks = compute_ranks(rows, order)

    from clean_workbook_sheets import model_columns

    cols = model_columns(ward)
    widths = [13, 13, 17] + ([20] if ward else [12]) + [14] * (len(cols) - 4)

    ws["A1"] = f"{'Ward' if ward else 'LGA'} model"
    ws["A1"].font = TITLE
    ws["C1"] = ("Static source data, then live calculations. "
                "Yellow inputs live on the Inputs sheet; nothing here needs editing.")
    ws["C1"].font = SMALL

    header_row(ws, 2, cols, widths, freeze=f"E3")
    first, last = 3, 2 + len(rows)

    # column letters, by header name
    C = {name_: gcl(i) for i, name_ in enumerate(cols, 1)}

    def col(name_):
        return C[name_]

    sam = col("U5 SAM (%)")
    need = col("Cartons needed")

    for n, row in enumerate(rows):
        r = first + n
        vals = [
            row["state"], zones.get(row["state"], ""), row["lga"],
            row.get("ward") if ward else row.get("nWards"),
            row["popTotal"], row["u5mr"], row["stunting"], row["wasting"],
            row["sam"], row["anemia"],
            ranks[n]["mortality"], ranks[n]["stunting"], ranks[n]["threshold"],
            "yes" if row.get("estimated") else "",
        ]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY
        ws.cell(row=r, column=cols.index("Population (total)") + 1).number_format = N0
        for nm in ("U5 mortality (per 1,000)", "U5 stunting (%)", "U5 wasting (%)",
                   "U5 SAM (%)", "U5 anemia (%)"):
            ws.cell(row=r, column=cols.index(nm) + 1).number_format = N1

        f = {
            "Population under 5": f"={col('Population (total)')}{r}*U5Share",
            "Population in age range":
                f"={col('Population under 5')}{r}*AgeShare*(1-{sam}{r}/100)",
            "Monthly cohort": f"={col('Population in age range')}{r}/AgeCohorts",
            "Population eligible":
                f"={col('Population in age range')}{r}+{col('Monthly cohort')}{r}*(EnrollPeriod-1)",
            "Cartons needed":
                f"=ROUNDUP({col('Population eligible')}{r}*SachetsPerChild/SachetsPerCarton*CoverageCap,0)",
            "Risk category": risk_formula(r, col("U5 mortality (per 1,000)"),
                                          col("U5 stunting (%)"), col("U5 wasting (%)")),
            "Annual U2 deaths":
                f"={col('U5 mortality (per 1,000)')}{r}/1000*{col('Population eligible')}{r}/5",
            "Stunting cases":
                f"={col('U5 stunting (%)')}{r}/100*{col('Population eligible')}{r}",
            "SAM cases": f"={sam}{r}/100*{col('Population eligible')}{r}",
            "Anemia cases":
                f"={col('U5 anemia (%)')}{r}/100*{col('Population eligible')}{r}",
            "Deaths averted": f"={col('Annual U2 deaths')}{r}*EffectMortality*EffectScale",
            "Stunting averted": f"={col('Stunting cases')}{r}*EffectStunting*EffectScale",
            "SAM averted": f"={col('SAM cases')}{r}*EffectSam*EffectScale",
            "Anemia averted": f"={col('Anemia cases')}{r}*EffectAnemia*EffectScale",
            "DALYs averted":
                f"={col('Deaths averted')}{r}*YllPerDeath+{col('SAM averted')}{r}*YldPerSam"
                f"+{col('Anemia averted')}{r}*YldPerAnemia",
            "Manual cartons for state":
                f"=IF(UseManual<>\"Yes\",0,IFERROR(INDEX(ManualCartons,MATCH(A{r},ManualStates,0)),0))",
            "Meets thresholds":
                f"=IF(UseThresholds<>\"Yes\",1,"
                f"IF(AND({col('U5 mortality (per 1,000)')}{r}>=ThreshU5mr,"
                f"{col('U5 stunting (%)')}{r}>=ThreshStunting,"
                f"{col('U5 wasting (%)')}{r}>=ThreshWasting),1,0))",
        }
        for nm, formula in f.items():
            ws.cell(row=r, column=cols.index(nm) + 1, value=formula).font = BODY

        # the four strategies
        for sname, rank_col in (("deaths", "Priority rank: deaths"),
                                ("stunting", "Priority rank: stunting"),
                                ("threshold", "Priority rank: threshold")):
            gate = f"{col('Meets thresholds')}{r}*" if sname == "threshold" else ""
            rk = col(rank_col)
            man = col(f"Manual: {sname}")
            rem = col(f"Remaining need: {sname}")
            alloc = col(f"ALLOCATED: {sname}")
            mancol = col("Manual cartons for state")

            # Manual fill: each state's reserved cartons, down that state's own
            # priority order. SUMIFS accumulates need strictly above this rank
            # within the same state, mirroring the original's growing window.
            ws[f"{man}{r}"] = (
                f"={gate}IF({mancol}{r}=0,0,"
                f"MIN({need}{r},MAX(0,{mancol}{r}"
                f"-SUMIFS(${need}${first}:${need}${last},$A${first}:$A${last},$A{r},"
                f"${rk}${first}:${rk}${last},\"<\"&${rk}{r}))))"
            )
            ws[f"{rem}{r}"] = f"={gate}({need}{r}-{man}{r})"
            # Pool fill, sort-free. See the module docstring for why cumulative
            # allocation above equals MIN(cumulative remaining need, pool).
            ws[f"{alloc}{r}"] = (
                f"={man}{r}+MIN({rem}{r},MAX(0,StrategyPool"
                f"-SUMIFS(${rem}${first}:${rem}${last},"
                f"${rk}${first}:${rk}${last},\"<\"&${rk}{r})))"
            )
            for cc in (man, rem, alloc):
                ws[f"{cc}{r}"].font = BODY
                ws[f"{cc}{r}"].number_format = N1

        # Equal distribution ignores manual allocation and thresholds entirely.
        # The national total sits in the cell directly above the header.
        ws[f"{col('ALLOCATED: equal')}{r}"] = (
            f"=IF({need}$1=0,0,TotalCartons*{need}{r}/{need}$1)"
        )
        a_d, a_s = col("ALLOCATED: deaths"), col("ALLOCATED: stunting")
        a_t, a_e = col("ALLOCATED: threshold"), col("ALLOCATED: equal")
        ws[f"{col('ALLOCATED: selected strategy')}{r}"] = (
            f"=IF(SelectedStrategy=StrategyName1,{a_d}{r},"
            f"IF(SelectedStrategy=StrategyName2,{a_s}{r},"
            f"IF(SelectedStrategy=StrategyName3,{a_t}{r},{a_e}{r})))"
        )
        for nm, fmt in (("Population under 5", N0), ("Population in age range", N0),
                        ("Monthly cohort", N1), ("Population eligible", N0),
                        ("Cartons needed", N0), ("Annual U2 deaths", N1),
                        ("Stunting cases", N0), ("SAM cases", N0), ("Anemia cases", N0),
                        ("Deaths averted", N2), ("Stunting averted", N1),
                        ("SAM averted", N1), ("Anemia averted", N1), ("DALYs averted", N1),
                        ("ALLOCATED: equal", N1), ("ALLOCATED: selected strategy", N1)):
            ws.cell(row=r, column=cols.index(nm) + 1).number_format = fmt

    # Column totals sit in row 1, directly above their own header, so the
    # equal-distribution share and the summary sheets have a single place to
    # read national figures from.
    for nm in ("Cartons needed", "ALLOCATED: deaths", "ALLOCATED: stunting",
               "ALLOCATED: threshold", "ALLOCATED: equal", "ALLOCATED: selected strategy"):
        letter = col(nm)
        c = ws[f"{letter}1"]
        c.value = f"=SUM({letter}{first}:{letter}{last})"
        c.font = BOLD
        c.number_format = N0
    ws.auto_filter.ref = f"A2:{gcl(len(cols))}{last}"
    ws.sheet_view.zoomScale = 90
    return ws, cols, first, last


def risk_formula(r, u5, st, wa):
    """Nested IF over the risk table on Assumptions.

    Level 1.3 deliberately omits the stunting test, reproducing the source
    defect so the rebuild agrees with the original. Flip `FixRiskBug` on the
    Inputs sheet to apply the intended rule instead.
    """
    parts = []
    for i, lvl in enumerate(("1.1", "1.2", "1.3", "2", "3")):
        row = 5 + i  # Assumptions rows for the risk table
        stunting = (
            f"OR(IF(FixRiskBug=\"Yes\",{st}{r}>=Risk_Stunting_{i},TRUE),{wa}{r}>=Risk_Wasting_{i})"
            if lvl == "1.3"
            else f"OR({st}{r}>=Risk_Stunting_{i},{wa}{r}>=Risk_Wasting_{i})"
        )
        parts.append(f"IF(AND({u5}{r}>=Risk_U5mr_{i},{stunting}),\"{lvl}\",")
    return "=" + "".join(parts) + '"Not Classified"' + ")" * 5


def main():
    if not os.path.exists(BASE):
        sys.exit(f"missing {BASE}; run tools/build_data.py first")
    with open(BASE, encoding="utf8") as fh:
        base = json.load(fh)
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    from clean_workbook_sheets import (
        build_readme, build_inputs, build_assumptions,
        build_outputs, build_comparison, build_quantification,
    )

    build_readme(wb, base)
    build_inputs(wb, base)
    build_assumptions(wb, base)
    build_model_sheet(wb, "Ward model", base["wards"], "wards", base["constants"], base["zones"])
    build_model_sheet(wb, "LGA model", base["lgas"], "lgas", base["constants"], base["zones"])
    build_outputs(wb, base)
    build_comparison(wb, base)
    build_quantification(wb, base)

    wb.save(OUT)
    size = os.path.getsize(OUT) / 1e6
    print(f"wrote {OUT}")
    print(f"  {len(wb.sheetnames)} sheets, all visible: {', '.join(wb.sheetnames)}")
    print(f"  {size:.2f} MB")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    main()
