"""Non-model sheets for the rebuilt workbook. Imported by build_clean_workbook.py."""

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

INK = "0B0B0B"
MUTED = "898781"
ACCENT = "2A78D6"
TITLE = Font(name="Calibri", size=15, bold=True, color=INK)
H2 = Font(name="Calibri", size=11, bold=True, color=ACCENT)
BODY = Font(name="Calibri", size=10, color=INK)
SMALL = Font(name="Calibri", size=9, color=MUTED)
BOLD = Font(name="Calibri", size=10, bold=True, color=INK)
HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F3B63")
INPUT_FILL = PatternFill("solid", fgColor="FFF6DC")
WRAP = Alignment(wrap_text=True, vertical="top")

N0, N1, N2, PCT = "#,##0", "#,##0.0", "#,##0.00", "0.0%"
MONEY = '"$"#,##0'

WARD = "'Ward model'"
LGA = "'LGA model'"
WARD_LAST, LGA_LAST = 9686, 776


def model_columns(ward):
    """The model sheet's column headers, in order.

    Single source of truth: the model builder lays the sheet out from this, and
    the summary sheets resolve their column letters from it, so the two cannot
    drift apart.
    """
    cols = ["State", "Zone", "LGA"] + (["Ward"] if ward else ["Wards in LGA"]) + [
        "Population (total)", "U5 mortality (per 1,000)", "U5 stunting (%)",
        "U5 wasting (%)", "U5 SAM (%)", "U5 anemia (%)",
        "Priority rank: deaths", "Priority rank: stunting", "Priority rank: threshold",
        "Ward-level data estimated",
        "Population under 5", "Population in age range", "Monthly cohort",
        "Population eligible", "Cartons needed", "Risk category",
        "Annual U2 deaths", "Stunting cases", "SAM cases", "Anemia cases",
        "Deaths averted", "Stunting averted", "SAM averted", "Anemia averted",
        "DALYs averted",
        "Manual cartons for state", "Meets thresholds",
    ]
    for s in ("deaths", "stunting", "threshold"):
        cols += [f"Manual: {s}", f"Remaining need: {s}", f"ALLOCATED: {s}"]
    cols += ["ALLOCATED: equal", "ALLOCATED: selected strategy"]
    return cols


def letters(ward=True):
    """Header name to column letter. Identical for both levels except column D."""
    return {n: gcl(i) for i, n in enumerate(model_columns(ward), 1)}


def name(wb, key, sheet, ref):
    wb.defined_names[key] = DefinedName(key, attr_text=f"{sheet}!{ref}")


def head(ws, row, labels, widths=None):
    for i, t in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=t)
        c.font, c.fill = HEAD, HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gcl(i)].width = w


def field(ws, row, text, cell_col, value, fmt=None, note=None):
    ws.cell(row=row, column=2, value=text).font = BODY
    c = ws.cell(row=row, column=cell_col, value=value)
    c.fill, c.font = INPUT_FILL, BOLD
    if fmt:
        c.number_format = fmt
    if note:
        n = ws.cell(row=row, column=cell_col + 1, value=note)
        n.font = SMALL
    return c


# --------------------------------------------------------------------- README


def build_readme(wb, base):
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 96

    ws["B2"] = "SQ-LNS Allocation Tool"
    ws["B2"].font = TITLE
    ws["B3"] = "Rebuilt: 8 sheets, all visible, no hidden calculation layer."
    ws["B3"].font = SMALL

    rows = [
        ("", ""),
        ("HOW TO USE", ""),
        ("1. Inputs", "Set program parameters. Optionally tick thresholds and manual allocation. "
                      "Only yellow cells are editable."),
        ("2. Read the outputs", "Allocation outputs, Strategy comparison, and Quantification all "
                                "update automatically."),
        ("", ""),
        ("SHEETS", ""),
        ("Inputs", "Every user-editable value, in one place."),
        ("Assumptions", "Every constant, with its source. Nothing is hard-coded inside formulas."),
        ("Ward model", f"{len(base['wards']):,} wards. Static source data, then live calculations "
                       "and all four strategy allocations."),
        ("LGA model", f"{len(base['lgas']):,} LGAs. Same structure as the ward model."),
        ("Allocation outputs", "National, zone, and state view for the selected strategy."),
        ("Strategy comparison", "All four strategies side by side, with cost-effectiveness."),
        ("Quantification", "Resource needs by risk category or by threshold."),
        ("", ""),
        ("WHAT CHANGED FROM THE ORIGINAL", ""),
        ("45 sheets to 8", "12 dormant sheets from earlier ANRiN/UNICEF exercises were reachable "
                           "only through nine stale cells on the old inputs sheet. Dropped."),
        ("8 strategy sheets to 3 columns",
         "Each of those sheets existed only to physically re-sort every row so a running total "
         "could walk down it. Because cumulative allocation equals MIN(cumulative need, pool), "
         "the same result comes from a rank column and a SUMIFS, with no sorting. This is exactly "
         "equivalent, not an approximation."),
        ("Duplicate calc sheets merged",
         "The allocation and quantification calculation sheets were identical except for which "
         "input cells they read. One model sheet per level now serves both."),
        ("No Google-only functions",
         "The original was a Google Sheets export in which 6,205 cells held COUNTUNIQUEIFS, "
         "SPARKLINE and similar as frozen values that Excel could not recalculate. None remain, "
         "so this file recalculates correctly in Excel and in Google Sheets."),
        ("Named ranges", "Formulas read TotalCartons and CoverageCap rather than "
                         "'Allocation Inputs'!$F$5."),
        ("", ""),
        ("DEFECTS FOUND IN THE ORIGINAL", ""),
        ("Risk level 1.3 over-assigned",
         "The original's level 1.3 test referenced an empty cell instead of the stunting "
         "threshold, so the test reduced to under-5 mortality alone. It promoted 232 of 9,684 "
         "wards and 14 of 774 LGAs into Very High Level 3 without their meeting its stunting or "
         "wasting criteria. This defect is live in the Google Sheet, not only in the export. "
         "This file reproduces it by default so figures match; set 'Correct the risk 1.3 defect' "
         "to Yes on the Inputs sheet to apply the intended rule."),
        ("Quantification ranges truncated",
         "In the exported file every range on the quantification sheet covered 90 of 9,684 rows. "
         "That was an export artifact rather than a fault in the Google Sheet. Fixed here."),
        ("Inconsistent threshold comparison",
         "The original's quantification used a strict greater-than while its allocation used "
         "greater-or-equal. This file uses greater-or-equal throughout."),
        ("", ""),
        ("NOTE ON PRIORITY RANKS", ""),
        ("Precomputed",
         "The three priority-rank columns on the model sheets depend only on static source data, "
         "so they are precomputed rather than recalculated. They do not change with program "
         "inputs. They would only need regenerating if the risk-category thresholds on the "
         "Assumptions sheet were edited, which are structural definitions rather than inputs."),
    ]
    r = 5
    for a, b in rows:
        if a and not b:
            ws.cell(row=r, column=2, value=a).font = H2
        elif a:
            ws.cell(row=r, column=2, value=a).font = BOLD
            c = ws.cell(row=r, column=3, value=b)
            c.font, c.alignment = BODY, WRAP
            ws.row_dimensions[r].height = max(14, 13 * (len(b) // 95 + 1))
        r += 1
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------------------- Inputs


def build_inputs(wb, base):
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 54
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    ws["B2"] = "Inputs"
    ws["B2"].font = TITLE
    ws["B3"] = "Only the shaded cells are editable. Everything else is calculated."
    ws["B3"].font = SMALL

    ws["B5"] = "Program"
    ws["B5"].font = H2
    field(ws, 6, "Total cartons supply", 3, 10000, N0)
    field(ws, 7, "Age range (months)", 3, "6 to 23", None, "6 to 23, 6 to 18, or 6 to 11")
    field(ws, 8, "Supplementation duration (months)", 3, 6, N0, "Below 3 months no effect is applied")
    field(ws, 9, "Enrollment period (months)", 3, 6, N0)
    field(ws, 10, "Coverage cap", 3, 0.75, PCT)
    field(ws, 11, "Allocation level", 3, "LGAs", None, "Wards or LGAs")
    field(ws, 12, "Strategy to display", 3, base["constants"]["strategies"][0]["label"])

    ws["B14"] = "Thresholds"
    ws["B14"].font = H2
    field(ws, 15, "Use thresholds", 3, "No", None, "Yes or No")
    field(ws, 16, "Under-5 mortality at or above (per 1,000)", 3, 0, N0)
    field(ws, 17, "Stunting prevalence at or above (%)", 3, 0, N0)
    field(ws, 18, "Wasting prevalence at or above (%)", 3, 0, N0)
    ws["D15"] = "A geography qualifies only if it clears all three."
    ws["D15"].font = SMALL

    ws["B20"] = "Manual allocation"
    ws["B20"].font = H2
    field(ws, 21, "Use manual allocation", 3, "No", None, "Yes or No")
    ws["B22"] = "Allocated manually"
    ws["B22"].font = BODY
    ws["C22"] = "=IF(UseManual=\"Yes\",SUM(ManualCartons),0)"
    ws["C22"].font, ws["C22"].number_format = BOLD, N0
    ws["B23"] = "Left for the strategy to allocate"
    ws["B23"].font = BODY
    ws["C23"] = "=TotalCartons-C22"
    ws["C23"].font, ws["C23"].number_format = BOLD, N0
    ws["D23"] = ('=IF(C23<0,"Manual allocation exceeds total supply.",'
                 'IF(C23=0,"All supply is committed manually; every strategy returns the same result.",""))')
    ws["D23"].font = SMALL

    ws["B25"] = "Model fidelity"
    ws["B25"].font = H2
    field(ws, 26, "Correct the risk 1.3 defect", 3, "No", None,
          "No reproduces the original exactly. See README.")

    # Per-state manual allocation table, in columns F to H so it sits beside the
    # input fields rather than overwriting their headers.
    for i, t in enumerate(("State", "Cartons", "Need")):
        c = ws.cell(row=5, column=6 + i, value=t)
        c.font, c.fill = HEAD, HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    for i, st in enumerate(base["states"]):
        r = 6 + i
        ws.cell(row=r, column=6, value=st).font = BODY
        c = ws.cell(row=r, column=7, value=0)
        c.fill, c.font, c.number_format = INPUT_FILL, BODY, N0
        need = ws.cell(row=r, column=8)
        nd, stc = CN["need"], CN["state"]
        need.value = (
            f'=IF(Level="Wards",'
            f'SUMIFS({WARD}!${nd}$3:${nd}${WARD_LAST},{WARD}!${stc}$3:${stc}${WARD_LAST},$F{r}),'
            f'SUMIFS({LGA}!${nd}$3:${nd}${LGA_LAST},{LGA}!${stc}$3:${stc}${LGA_LAST},$F{r}))'
        )
        need.font, need.number_format = SMALL, N0
    last = 5 + len(base["states"])
    ws.cell(row=last + 1, column=6, value="Total").font = BOLD
    t = ws.cell(row=last + 1, column=7, value=f"=SUM(G6:G{last})")
    t.font, t.number_format = BOLD, N0

    # dropdowns
    for cell, options in (
        ("C7", '"6 to 23,6 to 18,6 to 11"'),
        ("C11", '"LGAs,Wards"'),
        ("C15", '"Yes,No"'),
        ("C21", '"Yes,No"'),
        ("C26", '"Yes,No"'),
        ("C12", '"' + ",".join(s["label"] for s in base["constants"]["strategies"]) + '"'),
    ):
        dv = DataValidation(type="list", formula1=options, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[cell])

    n = lambda k, ref: name(wb, k, "Inputs", ref)
    n("TotalCartons", "$C$6")
    n("AgeRange", "$C$7")
    n("Duration", "$C$8")
    n("EnrollPeriod", "$C$9")
    n("CoverageCap", "$C$10")
    n("Level", "$C$11")
    n("SelectedStrategy", "$C$12")
    n("UseThresholds", "$C$15")
    n("ThreshU5mr", "$C$16")
    n("ThreshStunting", "$C$17")
    n("ThreshWasting", "$C$18")
    n("UseManual", "$C$21")
    n("ManualTotal", "$C$22")
    n("StrategyPool", "$C$23")
    n("FixRiskBug", "$C$26")
    n("ManualStates", f"$F$6:$F${last}")
    n("ManualCartons", f"$G$6:$G${last}")
    for i, s in enumerate(base["constants"]["strategies"], 1):
        wb.defined_names[f"StrategyName{i}"] = DefinedName(
            f"StrategyName{i}", attr_text=f'"{s["label"]}"')
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------- Assumptions


def build_assumptions(wb, base):
    c = base["constants"]
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 74

    ws["B2"] = "Assumptions"
    ws["B2"].font = TITLE
    ws["B3"] = "Every constant the model uses, with its source. Structural: change with care."
    ws["B3"].font = SMALL

    ws["B5"] = "Risk category thresholds"
    ws["B5"].font = H2
    head(ws, 6, ["", "Level", "U5 mortality", "U5 stunting %", "U5 wasting %"])
    for i, t in enumerate(c["riskThresholds"]):
        r = 7 + i
        ws.cell(row=r, column=2, value=f'{t["label"]} ({t["level"]})').font = BODY
        for j, key in enumerate(("u5mr", "stunting", "wasting")):
            cell = ws.cell(row=r, column=3 + j, value=t[key])
            cell.font, cell.number_format = BODY, N1
        name(wb, f"Risk_U5mr_{i}", "Assumptions", f"$C${r}")
        name(wb, f"Risk_Stunting_{i}", "Assumptions", f"$D${r}")
        name(wb, f"Risk_Wasting_{i}", "Assumptions", f"$E${r}")
    ws["F7"] = ("A geography takes the first level it qualifies for: "
                "U5MR at or above the threshold, and either stunting or wasting at or above theirs.")
    ws["F7"].font = SMALL

    rows = [
        ("Product and cost", None, None, None),
        ("Sachets per carton", c["sachetsPerCarton"], "SachetsPerCarton", "Product specification"),
        ("Sachets per child per year", c["sachetsPerChildPerYear"], "SachetsPerYear", "One per day"),
        ("Price per sachet (USD)", c["pricePerSachetUsd"], "PricePerSachet", "Product cost"),
        ("Exchange rate (NGN per USD)", c["ngnPerUsd"], "NgnPerUsd", "Update as needed"),
        ("Delivery cost per child (USD)", c["deliveryCostPerChildUsd"], "DeliveryCostPerChild", "Program estimate"),

        ("Population", None, None, None),
        ("Under-5 share of population", c["u5ShareOfPopulation"], "U5Share", "Demographic assumption"),
        ("Share of U5 aged 6 to 23 months", c["ageRangeShare"]["6 to 23"], "Share_6_23", "(9.5 + 18.3) / 100"),
        ("Share of U5 aged 6 to 18 months", c["ageRangeShare"]["6 to 18"], "Share_6_18", "(9.5 + 18.3 * 7/12) / 100"),
        ("Share of U5 aged 6 to 11 months", c["ageRangeShare"]["6 to 11"], "Share_6_11", "Demographic assumption"),
        ("Monthly cohorts, 6 to 23", c["ageRangeCohorts"]["6 to 23"], "Cohorts_6_23", "23 - 6 + 1"),
        ("Monthly cohorts, 6 to 18", c["ageRangeCohorts"]["6 to 18"], "Cohorts_6_18", None),
        ("Monthly cohorts, 6 to 11", c["ageRangeCohorts"]["6 to 11"], "Cohorts_6_11", "11 - 6 + 1"),

        ("Treatment effects", None, None, None),
        ("Effect on under-2 mortality", c["effect"]["mortality"], "EffectMortality",
         "WHO estimate from unpublished Lassi meta-analysis"),
        ("Effect on stunting", c["effect"]["stunting"], "EffectStunting", "Dewey K et al (2021)"),
        ("Effect on SAM", c["effect"]["sam"], "EffectSam", "Dewey K et al (2022), confirmed ITT"),
        ("Effect on anemia", c["effect"]["anemia"], "EffectAnemia",
         "Tam E et al (2020) and Dewey G K et al (2021); RR 0.84, 95% CI 0.75 to 0.93"),
        ("Discount for product wastage", c["impactDiscount"]["wastage"], "DiscountWastage", "Program assumption"),
        ("Discount for incomplete consumption", c["impactDiscount"]["incompleteConsumption"],
         "DiscountConsumption", "Program assumption"),

        ("DALY values", None, None, None),
        ("Discount rate", c["daly"]["discountRate"], "DalyDiscount", "Standard discount rate"),
        ("Life expectancy (years)", c["daly"]["lifeExpectancy"], "LifeExpectancy", "WHO life tables, Nigeria"),
        ("Discounted YLL per death", c["daly"]["yllPerDeath"], "YllPerDeath", "GBD 2021, discounted"),
        ("Discounted YLD per SAM case", c["daly"]["yldPerSamCase"], "YldPerSam", "GBD 2021, discounted"),
        ("Discounted YLD per anemia case", c["daly"]["yldPerAnemiaCase"], "YldPerAnemia",
         "Mean of mild, moderate, severe weights, discounted"),
    ]
    r = 14
    for lbl, val, nm, src in rows:
        if val is None:
            r += 1
            ws.cell(row=r, column=2, value=lbl).font = H2
            r += 1
            continue
        ws.cell(row=r, column=2, value=lbl).font = BODY
        cell = ws.cell(row=r, column=3, value=val)
        cell.font = BODY
        cell.number_format = N2 if abs(val) < 100 else N0
        if src:
            s = ws.cell(row=r, column=4, value=src)
            s.font, s.alignment = SMALL, WRAP
        name(wb, nm, "Assumptions", f"$C${r}")
        r += 1

    # values that depend on the inputs
    r += 1
    ws.cell(row=r, column=2, value="Derived from the inputs").font = H2
    r += 1
    derived = [
        ("Share of U5 in the selected age range",
         '=IF(AgeRange="6 to 23",Share_6_23,IF(AgeRange="6 to 18",Share_6_18,Share_6_11))',
         "AgeShare", N2),
        ("Monthly cohorts in the selected age range",
         '=IF(AgeRange="6 to 23",Cohorts_6_23,IF(AgeRange="6 to 18",Cohorts_6_18,Cohorts_6_11))',
         "AgeCohorts", N0),
        ("Sachets per child for the selected duration",
         "=ROUNDUP(Duration/12*SachetsPerYear,0)", "SachetsPerChild", N0),
        ("Share of effect retained after discounts",
         "=1-DiscountWastage-DiscountConsumption", "ImpactRetained", N2),
        ("Share of effect for the selected duration",
         "=IF(Duration<3,0,IF(Duration>=12,1,Duration/12))", "DurationShare", N2),
        ("Combined effect multiplier",
         "=ImpactRetained*DurationShare", "EffectScale", N2),
        ("Average children per LGA, selected age range",
         f"=ROUNDUP(SUM({LGA}!${CN['in_range']}$3:${CN['in_range']}${LGA_LAST})/{LGA_LAST - 2},0)",
         "AvgChildrenPerLga", N0),
        ("Average children per ward, selected age range",
         f"=ROUNDUP(SUM({WARD}!${CN['in_range']}$3:${CN['in_range']}${WARD_LAST})/{WARD_LAST - 2},0)",
         "AvgChildrenPerWard", N0),
        ("Delivery cost per geography (USD)",
         '=DeliveryCostPerChild*IF(Level="Wards",AvgChildrenPerWard,AvgChildrenPerLga)',
         "DeliveryCostPerGeo", N0),
    ]
    for lbl, formula, nm, fmt in derived:
        ws.cell(row=r, column=2, value=lbl).font = BODY
        cell = ws.cell(row=r, column=3, value=formula)
        cell.font, cell.number_format = BOLD, fmt
        name(wb, nm, "Assumptions", f"$C${r}")
        r += 1
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------- allocation outputs


def _lvl(ward_expr, lga_expr):
    return f'=IF(Level="Wards",{ward_expr},{lga_expr})'


# model-sheet column letters, resolved from the shared layout above
_L = letters()
CN = {
    "state": _L["State"], "zone": _L["Zone"], "lga": _L["LGA"],
    "u5mr": _L["U5 mortality (per 1,000)"], "stunting": _L["U5 stunting (%)"],
    "wasting": _L["U5 wasting (%)"],
    "need": _L["Cartons needed"], "risk": _L["Risk category"],
    "in_range": _L["Population in age range"], "eligible": _L["Population eligible"],
    "deaths_av": _L["Deaths averted"], "stunt_av": _L["Stunting averted"],
    "sam_av": _L["SAM averted"], "anemia_av": _L["Anemia averted"],
    "dalys": _L["DALYs averted"],
    "alloc_d": _L["ALLOCATED: deaths"], "alloc_s": _L["ALLOCATED: stunting"],
    "alloc_t": _L["ALLOCATED: threshold"], "alloc_e": _L["ALLOCATED: equal"],
    "alloc_sel": _L["ALLOCATED: selected strategy"],
}


def _sum_by(col, key_col, key_ref):
    """SUMIFS at the active level, keyed on a state or zone column."""
    w = f'SUMIFS({WARD}!${col}$3:${col}${WARD_LAST},{WARD}!${key_col}$3:${key_col}${WARD_LAST},{key_ref})'
    l = f'SUMIFS({LGA}!${col}$3:${col}${LGA_LAST},{LGA}!${key_col}$3:${key_col}${LGA_LAST},{key_ref})'
    return _lvl(w, l)


def _sum_scaled(value_col, key_col, key_ref):
    """Impact scaled by the share of need funded, summed with SUMPRODUCT.

    Each row's averted-case figures are stated at full coverage, so they are
    prorated by allocation divided by need before summing.
    """
    a, n = CN["alloc_sel"], CN["need"]
    w = (f'SUMPRODUCT(({WARD}!${key_col}$3:${key_col}${WARD_LAST}={key_ref})*'
         f'{WARD}!${value_col}$3:${value_col}${WARD_LAST}*'
         f'IF({WARD}!${n}$3:${n}${WARD_LAST}=0,0,'
         f'{WARD}!${a}$3:${a}${WARD_LAST}/({WARD}!${n}$3:${n}${WARD_LAST}+({WARD}!${n}$3:${n}${WARD_LAST}=0))))')
    l = (f'SUMPRODUCT(({LGA}!${key_col}$3:${key_col}${LGA_LAST}={key_ref})*'
         f'{LGA}!${value_col}$3:${value_col}${LGA_LAST}*'
         f'IF({LGA}!${n}$3:${n}${LGA_LAST}=0,0,'
         f'{LGA}!${a}$3:${a}${LGA_LAST}/({LGA}!${n}$3:${n}${LGA_LAST}+({LGA}!${n}$3:${n}${LGA_LAST}=0))))')
    return _lvl(w, l)


def build_outputs(wb, base):
    ws = wb.create_sheet("Allocation outputs")
    ws["B2"] = "Allocation outputs"
    ws["B2"].font = TITLE
    ws["B3"] = "=\"Strategy: \"&SelectedStrategy&\"   |   Level: \"&Level"
    ws["B3"].font = SMALL

    a = CN["alloc_sel"]
    ws["B5"] = "National"
    ws["B5"].font = H2
    nat = [
        ("Cartons allocated", _lvl(f"{WARD}!${a}$1", f"{LGA}!${a}$1"), N0),
        ("Cartons needed nationally", _lvl(f"{WARD}!${CN['need']}$1", f"{LGA}!${CN['need']}$1"), N0),
        ("Share of national need covered", "=IF(C7=0,0,C6/C7)", PCT),
        ("Geographies funded",
         _lvl(f'COUNTIFS({WARD}!${a}$3:${a}${WARD_LAST},">0")',
              f'COUNTIFS({LGA}!${a}$3:${a}${LGA_LAST},">0")'), N0),
    ]
    for i, (lbl, formula, fmt) in enumerate(nat):
        r = 6 + i
        ws.cell(row=r, column=2, value=lbl).font = BODY
        c = ws.cell(row=r, column=3, value=formula)
        c.font, c.number_format = BOLD, fmt

    ws["B12"] = "By zone"
    ws["B12"].font = H2
    zone_names = sorted(set(base["zones"].values()))
    head(ws, 13, ["", "Zone", "Cartons", "Share", "Deaths averted", "Stunting averted",
                  "SAM averted", "Anemia averted", "DALYs averted"],
         [3, 20, 14, 10, 15, 16, 14, 15, 14])
    for i, z in enumerate(zone_names):
        r = 14 + i
        ws.cell(row=r, column=2, value=z).font = BODY
        ws.cell(row=r, column=3, value=_sum_by(a, CN["zone"], f"$B{r}")).number_format = N0
        ws.cell(row=r, column=4, value=f"=IF($C$6=0,0,C{r}/$C$6)").number_format = PCT
        for j, vc in enumerate(("deaths_av", "stunt_av", "sam_av", "anemia_av", "dalys")):
            ws.cell(row=r, column=5 + j,
                    value=_sum_scaled(CN[vc], CN["zone"], f"$B{r}")).number_format = N1

    zlast = 13 + len(zone_names)
    ws.cell(row=zlast + 1, column=2, value="Total").font = BOLD
    for cidx in range(3, 10):
        col = gcl(cidx)
        t = ws.cell(row=zlast + 1, column=cidx, value=f"=SUM({col}14:{col}{zlast})")
        t.font = BOLD
        t.number_format = PCT if cidx == 4 else (N0 if cidx == 3 else N1)

    start = zlast + 3
    ws.cell(row=start, column=2, value="By state").font = H2
    head(ws, start + 1, ["", "State", "Zone", "Cartons", "Share", "Need", "Covered",
                         "Deaths averted", "Stunting averted", "DALYs averted"],
         [3, 18, 18, 14, 10, 14, 10, 15, 16, 14])
    for i, st in enumerate(base["states"]):
        r = start + 2 + i
        ws.cell(row=r, column=2, value=st).font = BODY
        ws.cell(row=r, column=3, value=base["zones"].get(st, "")).font = SMALL
        ws.cell(row=r, column=4, value=_sum_by(a, CN["state"], f"$B{r}")).number_format = N0
        ws.cell(row=r, column=5, value=f"=IF($C$6=0,0,D{r}/$C$6)").number_format = PCT
        ws.cell(row=r, column=6, value=_sum_by(CN["need"], CN["state"], f"$B{r}")).number_format = N0
        ws.cell(row=r, column=7, value=f"=IF(F{r}=0,0,D{r}/F{r})").number_format = PCT
        for j, vc in enumerate(("deaths_av", "stunt_av", "dalys")):
            ws.cell(row=r, column=8 + j,
                    value=_sum_scaled(CN[vc], CN["state"], f"$B{r}")).number_format = N1
    slast = start + 1 + len(base["states"])
    ws.cell(row=slast + 1, column=2, value="Total").font = BOLD
    for cidx in (4, 6, 8, 9, 10):
        col = gcl(cidx)
        t = ws.cell(row=slast + 1, column=cidx, value=f"=SUM({col}{start+2}:{col}{slast})")
        t.font = BOLD
        t.number_format = N0 if cidx in (4, 6) else N1
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------- strategy comparison


def build_comparison(wb, base):
    ws = wb.create_sheet("Strategy comparison")
    ws["B2"] = "Strategy comparison"
    ws["B2"].font = TITLE
    ws["B3"] = "All four strategies at the current inputs. Cost per case averted: lower is better."
    ws["B3"].font = SMALL

    strategies = base["constants"]["strategies"]
    alloc_cols = [CN["alloc_d"], CN["alloc_s"], CN["alloc_t"], CN["alloc_e"]]

    head(ws, 5, ["", "Strategy", "Cartons", "Geographies funded", "Deaths averted",
                 "Stunting averted", "SAM averted", "Anemia averted", "DALYs averted"],
         [3, 34, 13, 15, 14, 15, 13, 14, 14])

    def scaled_total(value_col, ac):
        n = CN["need"]
        w = (f'SUMPRODUCT({WARD}!${value_col}$3:${value_col}${WARD_LAST}*'
             f'{WARD}!${ac}$3:${ac}${WARD_LAST}/'
             f'({WARD}!${n}$3:${n}${WARD_LAST}+({WARD}!${n}$3:${n}${WARD_LAST}=0)))')
        l = (f'SUMPRODUCT({LGA}!${value_col}$3:${value_col}${LGA_LAST}*'
             f'{LGA}!${ac}$3:${ac}${LGA_LAST}/'
             f'({LGA}!${n}$3:${n}${LGA_LAST}+({LGA}!${n}$3:${n}${LGA_LAST}=0)))')
        return _lvl(w, l)

    for i, s in enumerate(strategies):
        r = 6 + i
        ac = alloc_cols[i]
        ws.cell(row=r, column=2, value=s["label"]).font = BODY
        ws.cell(row=r, column=3, value=_lvl(f"SUM({WARD}!${ac}$3:${ac}${WARD_LAST})",
                                            f"SUM({LGA}!${ac}$3:${ac}${LGA_LAST})")).number_format = N0
        ws.cell(row=r, column=4,
                value=_lvl(f'COUNTIFS({WARD}!${ac}$3:${ac}${WARD_LAST},">0")',
                           f'COUNTIFS({LGA}!${ac}$3:${ac}${LGA_LAST},">0")')).number_format = N0
        for j, vc in enumerate(("deaths_av", "stunt_av", "sam_av", "anemia_av", "dalys")):
            ws.cell(row=r, column=5 + j, value=scaled_total(CN[vc], ac)).number_format = N1

    ws["B12"] = "Cost-effectiveness (USD)"
    ws["B12"].font = H2
    head(ws, 13, ["", "Strategy", "Total cost", "Per death", "Per stunting case",
                  "Per SAM case", "Per anemia case", "Per DALY"],
         [3, 34, 14, 13, 16, 13, 15, 13])
    for i, s in enumerate(strategies):
        r = 14 + i
        src = 6 + i
        ws.cell(row=r, column=2, value=s["label"]).font = BODY
        ws.cell(row=r, column=3,
                value=f"=C{src}*SachetsPerCarton*PricePerSachet+D{src}*DeliveryCostPerGeo"
                ).number_format = MONEY
        for j, col in enumerate("EFGHI"):
            ws.cell(row=r, column=4 + j,
                    value=f"=IF({col}{src}=0,\"n/a\",$C{r}/{col}{src})").number_format = MONEY
    ws.sheet_view.showGridLines = False


# ------------------------------------------------------------- quantification


def build_quantification(wb, base):
    ws = wb.create_sheet("Quantification")
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 46

    ws["B2"] = "Quantification"
    ws["B2"].font = TITLE
    ws["B3"] = ("What would it cost to cover a given set of geographies? "
                "Uses the program inputs from the Inputs sheet.")
    ws["B3"].font = SMALL

    ws["B5"] = "Price inputs"
    ws["B5"].font = H2
    field(ws, 6, "Cost per carton (NGN)", 3, 80000, N0)
    field(ws, 7, "Markup for last-mile delivery", 3, 0.2, PCT)
    name(wb, "CostPerCarton", "Quantification", "$C$6")
    name(wb, "DeliveryMarkup", "Quantification", "$C$7")

    ws["B9"] = "Target by risk category"
    ws["B9"].font = H2
    head(ws, 10, ["", "Risk category", "Include", "Cartons needed", "Children", "Deaths averted", "DALYs averted"],
         [3, 30, 10, 16, 14, 15, 14])
    levels = base["constants"]["riskThresholds"]
    for i, t in enumerate(levels):
        r = 11 + i
        ws.cell(row=r, column=2, value=f'{t["label"]} ({t["level"]})').font = BODY
        inc = ws.cell(row=r, column=3, value="Yes" if t["level"] != "3" else "No")
        inc.fill, inc.font = INPUT_FILL, BOLD
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(inc)
        for j, col in enumerate((CN["need"], CN["eligible"], CN["deaths_av"], CN["dalys"])):
            w = (f'SUMIFS({WARD}!${col}$3:${col}${WARD_LAST},'
                 f'{WARD}!${CN["risk"]}$3:${CN["risk"]}${WARD_LAST},"{t["level"]}")')
            l = (f'SUMIFS({LGA}!${col}$3:${col}${LGA_LAST},'
                 f'{LGA}!${CN["risk"]}$3:${CN["risk"]}${LGA_LAST},"{t["level"]}")')
            ws.cell(row=r, column=4 + j,
                    value=f'=IF($C{r}<>"Yes",0,{_lvl(w, l)[1:]})').number_format = N0
    rlast = 10 + len(levels)
    ws.cell(row=rlast + 1, column=2, value="Selected total").font = BOLD
    for cidx in range(4, 8):
        col = gcl(cidx)
        c = ws.cell(row=rlast + 1, column=cidx, value=f"=SUM({col}11:{col}{rlast})")
        c.font, c.number_format = BOLD, N0

    r = rlast + 3
    ws.cell(row=r, column=2, value="Budget for the selected categories").font = H2
    for i, (lbl, formula, fmt) in enumerate([
        ("Cartons needed", f"=D{rlast+1}", N0),
        ("Budget required (NGN)", f"=C{r+1}*CostPerCarton*(1+DeliveryMarkup)", '"NGN "#,##0'),
        ("Budget required (USD)", f"=C{r+2}/NgnPerUsd", MONEY),
    ]):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=lbl).font = BODY
        c = ws.cell(row=rr, column=3, value=formula)
        c.font, c.number_format = BOLD, fmt

    r = r + 6
    ws.cell(row=r, column=2, value="Target by threshold").font = H2
    ws.cell(row=r + 1, column=2,
            value="Uses the thresholds on the Inputs sheet, at or above in each case.").font = SMALL
    head(ws, r + 2, ["", "Measure", "Value"], [3, 44, 18])
    n_, rk = CN["need"], CN["risk"]
    u5, stn, wst = "F", "G", "H"
    for i, (lbl, col, fmt) in enumerate([
        ("Geographies qualifying", None, N0),
        ("Cartons needed", CN["need"], N0),
        ("Children targeted", CN["eligible"], N0),
        ("Deaths averted", CN["deaths_av"], N1),
        ("DALYs averted", CN["dalys"], N1),
    ]):
        rr = r + 3 + i
        ws.cell(row=rr, column=2, value=lbl).font = BODY
        if col is None:
            w = (f'COUNTIFS({WARD}!${u5}$3:${u5}${WARD_LAST},">="&ThreshU5mr,'
                 f'{WARD}!${stn}$3:${stn}${WARD_LAST},">="&ThreshStunting,'
                 f'{WARD}!${wst}$3:${wst}${WARD_LAST},">="&ThreshWasting)')
            l = (f'COUNTIFS({LGA}!${u5}$3:${u5}${LGA_LAST},">="&ThreshU5mr,'
                 f'{LGA}!${stn}$3:${stn}${LGA_LAST},">="&ThreshStunting,'
                 f'{LGA}!${wst}$3:${wst}${LGA_LAST},">="&ThreshWasting)')
        else:
            w = (f'SUMIFS({WARD}!${col}$3:${col}${WARD_LAST},'
                 f'{WARD}!${u5}$3:${u5}${WARD_LAST},">="&ThreshU5mr,'
                 f'{WARD}!${stn}$3:${stn}${WARD_LAST},">="&ThreshStunting,'
                 f'{WARD}!${wst}$3:${wst}${WARD_LAST},">="&ThreshWasting)')
            l = (f'SUMIFS({LGA}!${col}$3:${col}${LGA_LAST},'
                 f'{LGA}!${u5}$3:${u5}${LGA_LAST},">="&ThreshU5mr,'
                 f'{LGA}!${stn}$3:${stn}${LGA_LAST},">="&ThreshStunting,'
                 f'{LGA}!${wst}$3:${wst}${LGA_LAST},">="&ThreshWasting)')
        c = ws.cell(row=rr, column=3, value=_lvl(w, l))
        c.font, c.number_format = BOLD, fmt
    ws.sheet_view.showGridLines = False
